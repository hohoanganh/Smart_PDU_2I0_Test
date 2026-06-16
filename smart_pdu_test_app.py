#!/usr/bin/env python3
"""
SMART PDU 2.0 - Test App (GUI)
==============================
App desktop test Smart PDU 2.0 (STM32L151) qua CLI serial (USART1, 115200).

Chay truc tiep:   pip install pyserial openpyxl   ->   python smart_pdu_test_app.py
Build .exe:       chay build_pdu_exe.bat (can pyinstaller)
"""

import os
import queue
import re
import sys
import threading
import time
import tkinter as tk
from tkinter import ttk, messagebox

try:
    import serial
    import serial.tools.list_ports
except ImportError:
    print("Thieu pyserial. Cai bang:  pip install pyserial")
    sys.exit(1)

APP_TITLE   = "Smart PDU 2.0 - App Test"
DEVICE_ID_SIG = "SMART_PDU_2I0"   # ID xác thực dùng thiết bị (lệnh 'id')
BAUD        = 115200
BAUDRATES   = [1200, 2400, 4800, 9600, 19200, 38400, 57600, 115200,
               230400, 460800, 921600]
CMD_TIMEOUT = 3.0
N_RELAY     = 4
RL_SEQ_DELAY = 0.35   # delay giữa mỗi relay khi ALL ON/OFF (tránh đóng dồn)

IMG_NAME    = "PDU_device.png"
FPT_LOGO    = "FPT_logo.png"
HDR_LOGO_H  = 56   # chiều cao chung cho 2 logo trên header (cho bằng nhau)

# ===== Palette =====
HDR_BG      = "#1b2a38"
HDR_ACC     = "#e8833a"
MAIN_BG     = "#f0f2f5"
CARD_BG     = "#f8fafc"
CARD_BD     = "#e2e8f0"
WHITE       = "white"
BLUE_ACC    = "#1e40af"
BLUE_BTN    = "#3b82f6"
BLUE_HOV    = "#2563eb"

RL_ON_BG    = "#16a34a"
RL_OFF_BG   = "#475569"

LED_ACTIVE_LOW = False  # FW đã xử lý active-low (LED_ON_LVL=LOW): "LEDn: ON" =
                        # LED THẬT SỰ SÁNG. App hiển thị trực tiếp, không đảo.
                        # (Cần nạp lại firmware mới để đồng bộ đúng.)
LED_ON_FILL  = "#22c55e"
LED_OFF_FILL = "#334155"
LED_ON_LINE  = "#15803d"
LED_OFF_LINE = "#64748b"

BTN_DOWN_FILL = "#3b82f6"
BTN_UP_FILL   = "#334155"
BTN_DOWN_LINE = "#1d4ed8"
BTN_UP_LINE   = "#64748b"

RX_RELAY_RE      = re.compile(r"RL([1-3]|_BATT)\s*:\s*(ON|OFF)")
RX_LED_RE        = re.compile(r"LED([1-4])\s*:\s*(ON|OFF)")
RX_BTN_RE        = re.compile(r"BT([1-4])\s*:\s*(DOWN|UP)")
RX_FW_RESET_RE   = re.compile(r"SYSTEM INIT")
RX_BTN_ERR_HOLD  = re.compile(r"BT_ERR:\s*HOLD\s*BT([1-4])")
RX_BTN_ERR_MULTI = re.compile(r"BT_ERR:\s*MULTI")
RX_DIP_RE        = re.compile(r"DIP:\s*0x([0-9A-Fa-f]+)")
RX_RS485_BAUD_RE = re.compile(r"RS485 BAUD:\s*(\d+)")


def resource_path(name):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


REPORT_NAME = "test_report.xlsx"

def report_path():
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, REPORT_NAME)


# ===== Test sequence =====
# Mỗi bước có "kind":
#   auto   - gửi lệnh, kiểm chuỗi must/must_not (tự động PASS/FAIL)
#   led    - bật LED 1->4 lần lượt, người dùng xác nhận (PASS) / Bỏ qua (SKIP)
#   button - người dùng nhấn BT1->4 (đủ 4 nút = PASS) / Bỏ qua (SKIP)
#   rs485  - gửi text qua RS485, đợi nhận lại đúng text (loopback) / Bỏ qua
RS485_TEST_TEXT = "Smart PDU RS485 Test"
TESTS = [
    {"kind": "auto",   "name": "i2c scan (PCA @ 0x38)", "cmd": "i2c",     "wait": 0.6,
     "must": ["I2C: 0x38", "OK"], "must_not": []},
    {"kind": "auto",   "name": "Flash ID/Read/RW",    "cmd": "fwr",      "wait": 1.5,
     "must": ["FLASH OK"],       "must_not": ["FLASH FAIL"]},
    {"kind": "auto",   "name": "Relay 1 ON",  "cmd": "rl 1 on",  "wait": 0.5,
     "must": ["RL1: ON",  "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay 1 OFF", "cmd": "rl 1 off", "wait": 0.5,
     "must": ["RL1: OFF", "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay 2 ON",  "cmd": "rl 2 on",  "wait": 0.5,
     "must": ["RL2: ON",  "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay 2 OFF", "cmd": "rl 2 off", "wait": 0.5,
     "must": ["RL2: OFF", "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay 3 ON",  "cmd": "rl 3 on",  "wait": 0.5,
     "must": ["RL3: ON",  "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay 3 OFF", "cmd": "rl 3 off", "wait": 0.5,
     "must": ["RL3: OFF", "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay BATT ON",  "cmd": "rl 4 on",  "wait": 0.5,
     "must": ["RL_BATT: ON",  "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "auto",   "name": "Relay BATT OFF", "cmd": "rl 4 off", "wait": 0.5,
     "must": ["RL_BATT: OFF", "OK"], "must_not": ["PCA FAIL"]},
    {"kind": "led",    "name": "LED 1-4 sáng lần lượt (xác nhận)"},
    {"kind": "button", "name": "Nút nhấn BT1-4 (nhấn từng nút)"},
    {"kind": "dip",    "name": "Dip Switch 1-4 (gạt từng bit)"},
    {"kind": "rs485",  "name": "RS485 gửi/nhận loopback", "text": RS485_TEST_TEXT},
]


# =============================================================
#  SerialWorker
# =============================================================
class SerialWorker:
    def __init__(self, log_queue):
        self.ser        = None
        self.log_queue  = log_queue
        self._rx_buf    = b""
        self._rx_lock   = threading.Lock()
        self._cmd_lock  = threading.Lock()   # tránh 2 send_cmd chen nhau
        self._stop      = threading.Event()
        self._thread    = None

    def open(self, port, baud=BAUD):
        self.ser = serial.Serial(port, baud, timeout=0.1)
        self._stop.clear()
        self._thread = threading.Thread(target=self._reader, daemon=True)
        self._thread.start()

    def close(self):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=1.0)
            self._thread = None
        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass
            self.ser = None

    @property
    def is_open(self):
        return self.ser is not None and self.ser.is_open

    def _reader(self):
        while not self._stop.is_set():
            try:
                chunk = self.ser.read(self.ser.in_waiting or 1)
            except Exception:
                self.log_queue.put(("err", "\n[Mất kết nối serial]\n"))
                break
            if chunk:
                with self._rx_lock:
                    self._rx_buf += chunk
                self.log_queue.put(("rx", chunk.decode(errors="replace")))

    def clear_rx(self):
        with self._rx_lock:
            self._rx_buf = b""
        if self.ser:
            try:
                self.ser.reset_input_buffer()
            except Exception:
                pass

    def take_rx(self):
        with self._rx_lock:
            data, self._rx_buf = self._rx_buf, b""
        return data

    def write_line(self, text):
        if self.is_open:
            self.ser.write((text + "\n").encode())
            self.log_queue.put(("tx", f"> {text}\n"))

    def send_cmd(self, cmd, wait=0.5):
        # Khóa: đảm bảo 1 lệnh hoàn tất (clear_rx -> write -> đọc hết) trước khi
        # lệnh khác chạy -> tránh _query_fw_ver và test 'ver' đọc lần phản hồi.
        with self._cmd_lock:
            self.clear_rx()
            self.write_line(cmd)
            time.sleep(wait)
            deadline = time.time() + CMD_TIMEOUT
            out = b""
            while time.time() < deadline:
                chunk = self.take_rx()
                if chunk:
                    out += chunk
                    deadline = time.time() + 0.3
                else:
                    time.sleep(0.05)
                    if not self._rx_buf:
                        break
            out += self.take_rx()
            return out.decode(errors="replace")


# =============================================================
#  Helper
# =============================================================
def _make_dot(parent, size=10, bg=WHITE, color="#94a3b8"):
    c   = tk.Canvas(parent, width=size, height=size, bg=bg, highlightthickness=0)
    oid = c.create_oval(1, 1, size-1, size-1, fill=color, outline="")
    return c, oid


# =============================================================
#  App
# =============================================================
class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1300x780")
        self.configure(bg=MAIN_BG)

        self.log_queue      = queue.Queue()
        self.worker         = SerialWorker(self.log_queue)
        self.testing        = False
        self._stop_test     = threading.Event()
        self.relay_state    = [False] * (N_RELAY + 1)
        self.led_state      = [False] * (N_RELAY + 1)
        self._rl_busy_until = 0.0
        self._rx_scan_buf   = ""
        self._btn_down_time = {}
        self._btn_pre_relay = {}
        self._bt_test_active  = False   # đang chạy bước test nút nhấn?
        self._bt_test_pressed = set()   # các nút đã nhấn trong bước test
        self._dip_test_active = False   # đang chạy bước test Dip Switch?
        self._dip_test_seen   = set()   # các bit DIP đã gạt trong bước test
        self._dip_last_value  = None    # giá trị DIP gần nhất (để so sánh thay đổi)
        self._buzzer_muted    = False   # tắt tiếng bíp (test im lặng)
        self._cmd_history     = []      # lịch sử lệnh terminal (tối đa 50)
        self._cmd_hist_idx    = -1      # vị trí duyệt lịch sử (-1 = hiện tại)
        self._connecting        = False  # đang mở cổng / xác thực ID
        self._connect_cancelled = False  # người dùng đã bấm Hủy
        self._connect_timeout_id = None  # id self.after() của watchdog cảnh báo

        self._build_ui()
        self._refresh_ports()
        # Khóa kích thước tối thiểu = kích thước thật sự của nội dung.
        # Kéo nhỏ cửa sổ cũng không thể cắt/mất nội dung các thẻ; kéo to
        # thì các thẻ tự giãn ra (expand). Không dùng canvas cuộn để tránh
        # lỗi remap làm mất nội dung khi kéo resize.
        self.update_idletasks()
        self.minsize(self.winfo_reqwidth(), self.winfo_reqheight())
        self.after(50, self._poll_log)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ----------------------------------------------------------
    #  UI helpers
    # ----------------------------------------------------------
    def _card(self, parent, title, **pack_kw):
        frm = tk.Frame(parent, bg=WHITE, relief="raised", bd=2)
        frm.pack(**pack_kw)
        hf = tk.Frame(frm, bg=WHITE, padx=12, pady=8)
        hf.pack(fill="x")
        tk.Label(hf, text=title, bg=WHITE, fg=BLUE_ACC,
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Frame(frm, bg=CARD_BD, height=1).pack(fill="x")
        body = tk.Frame(frm, bg=WHITE)
        body.pack(fill="both", expand=True)
        return body

    def _flat_btn(self, parent, text, bg, fg="white", hover=None,
                  font_size=10, bold=True, raised=False, **kw):
        hover = hover or bg
        return tk.Button(parent, text=text, bg=bg, fg=fg,
                         activebackground=hover, activeforeground=fg,
                         disabledforeground="#94a3b8",
                         font=("Segoe UI", font_size, "bold" if bold else "normal"),
                         relief="raised" if raised else "flat",
                         bd=2 if raised else 0, **kw)

    # ----------------------------------------------------------
    #  BUILD UI
    # ----------------------------------------------------------
    def _build_ui(self):

        # ── HEADER ───────────────────────────────────────────
        hdr = tk.Frame(self, bg=HDR_BG)
        hdr.pack(fill="x")

        # Logo thiết bị (PDU_device) - resize về cùng chiều cao với logo FPT
        self._logo = None
        try:
            from PIL import Image as _PIL_Image, ImageTk as _PIL_ImageTk
            _im = _PIL_Image.open(resource_path(IMG_NAME)).convert("RGBA")
            _w, _h = _im.size
            _nw = max(1, int(_w * HDR_LOGO_H / _h))
            _im = _im.resize((_nw, HDR_LOGO_H), _PIL_Image.LANCZOS)
            _bgc = tuple(int(HDR_BG.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
            _bg = _PIL_Image.new("RGB", (_nw, HDR_LOGO_H), _bgc)
            _bg.paste(_im, mask=_im.split()[3])
            self._logo = _PIL_ImageTk.PhotoImage(_bg)
            tk.Label(hdr, image=self._logo, bg=HDR_BG).pack(
                side="left", padx=(12, 8), pady=8)
        except Exception as _e:
            # Fallback không có PIL: subsample theo chiều cao
            try:
                img = tk.PhotoImage(file=resource_path(IMG_NAME))
                f = max(1, img.height() // HDR_LOGO_H)
                self._logo = img.subsample(f, f)
                tk.Label(hdr, image=self._logo, bg=HDR_BG).pack(
                    side="left", padx=(12, 8), pady=8)
            except Exception:
                pass

        tf = tk.Frame(hdr, bg=HDR_BG)
        tf.pack(side="left", pady=10)
        tk.Label(tf, text="Smart PDU 2.0 - App Test",
                 bg=HDR_BG, fg=WHITE,
                 font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(tf, text="4x Latching Relay  *  PCA9554 @ 0x38  *  STM32L151",
                 bg=HDR_BG, fg=HDR_ACC,
                 font=("Segoe UI", 9)).pack(anchor="w")

        # FPT Telecom logo (bên phải title, trước controls) -- dùng PIL ImageTk
        self._fpt_logo = None
        try:
            from PIL import Image as _PIL_Image, ImageTk as _PIL_ImageTk
            _fpt_h_target = HDR_LOGO_H
            _img = _PIL_Image.open(resource_path(FPT_LOGO)).convert("RGBA")
            _w, _h = _img.size
            _new_w = int(_w * _fpt_h_target / _h)
            _img = _img.resize((_new_w, _fpt_h_target), _PIL_Image.LANCZOS)
            # Paste RGBA onto header background color to avoid white box
            _bg_color = tuple(int(HDR_BG.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
            _bg = _PIL_Image.new("RGB", (_new_w, _fpt_h_target), _bg_color)
            _bg.paste(_img, mask=_img.split()[3])
            self._fpt_logo = _PIL_ImageTk.PhotoImage(_bg)
            tk.Label(hdr, image=self._fpt_logo, bg=HDR_BG).pack(
                side="left", padx=(18, 4), pady=6)
        except Exception as _e:
            print(f"[FPT logo load failed: {_e}]")

        # ── CONTROL BAR (dưới header: COM / trạng thái / Run TEST) ──
        ctrl = tk.Frame(self, bg="#22384a", padx=12, pady=8)
        ctrl.pack(fill="x")
        tk.Frame(self, bg="#0e1822", height=1).pack(fill="x")

        CTRL_BG = "#22384a"

        # Trạng thái kết nối (bên phải)
        statf = tk.Frame(ctrl, bg=CTRL_BG)
        statf.pack(side="right")
        self._conn_dot, self._conn_dot_oid = _make_dot(statf, 12, CTRL_BG, "#64748b")
        self._conn_dot.pack(side="left", padx=(0, 5))
        self._lbl_conn = tk.Label(statf, text="Chưa kết nối",
                                   bg=CTRL_BG, fg="#94a3b8",
                                   font=("Segoe UI", 9, "bold"),
                                   width=16, anchor="w")   # cố định -> không nháy
        self._lbl_conn.pack(side="left")

        # COM port
        tk.Label(ctrl, text="COM:", bg=CTRL_BG, fg="#94a3b8",
                 font=("Segoe UI", 9)).pack(side="left")
        # Bề rộng cố định + KHÔNG expand -> khi đổi trạng thái kết nối, các
        # control không bị co giãn/nhảy ngang (giao diện đứng yên).
        self.cbo_port = ttk.Combobox(ctrl, width=46, state="readonly")
        self.cbo_port.pack(side="left", padx=(4, 2))
        self.btn_refresh_port = tk.Button(
            ctrl, text="↻", bg="#2d4a62", fg="white",
            activebackground="#3a5f7a", relief="raised", bd=2,
            font=("Segoe UI", 9, "bold"), width=2, pady=2,
            command=self._refresh_ports)
        self.btn_refresh_port.pack(side="left", padx=(0, 10))

        tk.Label(ctrl, text="Baud:", bg=CTRL_BG, fg="#94a3b8",
                 font=("Segoe UI", 9)).pack(side="left")
        self.cbo_baud = ttk.Combobox(ctrl, width=8, state="readonly",
                                      values=[str(b) for b in BAUDRATES])
        self.cbo_baud.set(str(BAUD))
        self.cbo_baud.pack(side="left", padx=(4, 2))
        tk.Label(ctrl, text="(Thiết bị: 115200)", bg=CTRL_BG, fg="#f59e0b",
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 10))

        self.btn_conn = self._flat_btn(ctrl, "Kết Nối", BLUE_BTN, hover=BLUE_HOV,
                                        font_size=9, padx=14, pady=4, width=14,
                                        raised=True, command=self._toggle_conn)
        self.btn_conn.pack(side="left", padx=(0, 10))

        tk.Label(ctrl, text="Serial:", bg=CTRL_BG, fg="#94a3b8",
                 font=("Segoe UI", 9)).pack(side="left")
        self.ent_serial = ttk.Entry(ctrl, width=14)
        self.ent_serial.pack(side="left", padx=(4, 8))

        self.btn_run = self._flat_btn(ctrl, "Run Test", BLUE_BTN, hover=BLUE_HOV,
                                       font_size=9, padx=14, pady=4, width=14,
                                       raised=True, state="disabled", command=self._run_tests)
        self.btn_run.pack(side="left")

        # ── DEVICE INFO BAR ──────────────────────────────────
        info_bar = tk.Frame(self, bg=WHITE, pady=2,
                             highlightbackground=CARD_BD, highlightthickness=1)
        info_bar.pack(fill="x")
        self._dev_info = {}
        for _k, _v, _w in [("Firmware:", "--", 10),
                            ("MCU:", "STM32L151", 12),
                            ("Baud:", str(BAUD), 8)]:
            tk.Label(info_bar, text=_k, bg=WHITE, fg="#6b7280",
                     font=("Segoe UI", 9), padx=8, pady=3).pack(side="left")
            lv = tk.Label(info_bar, text=_v, bg=WHITE, fg="#1e293b",
                           font=("Segoe UI", 9, "bold"), width=_w, anchor="w")
            lv.pack(side="left")
            self._dev_info[_k] = lv
            tk.Label(info_bar, text="|", bg=WHITE, fg="#cbd5e1",
                     font=("Segoe UI", 9)).pack(side="left", padx=6)

        #  Ghi chú baud chuẩn của thiết bị (tránh chọn đúng COM nhưng sai baud)
        tk.Label(info_bar, text="Baud thiết bị: Console 115200",
                 bg=WHITE, fg="#e8833a", font=("Segoe UI", 9, "bold"),
                 padx=4, pady=3).pack(side="left")

        # ── BODY ─────────────────────────────────────────────
        body = tk.Frame(self, bg=MAIN_BG)
        body.pack(fill="both", expand=True)

        content = tk.Frame(body, bg=MAIN_BG)
        content.pack(fill="both", expand=True)

        # ── TOP ROW ──────────────────────────────────────────
        top_row = tk.Frame(content, bg=MAIN_BG)
        top_row.pack(fill="x", padx=10, pady=(10, 8))

        # ── RELAY CHANNEL CARDS ──────────────────────────────
        rc_body = self._card(top_row, "Điều Khiển Relay",
                              side="left", fill="both", expand=True, padx=(0, 8))
        ch_row = tk.Frame(rc_body, bg=WHITE, padx=10, pady=10)
        ch_row.pack(expand=True)

        self.btn_relay  = {}
        self.led_ind    = {}
        self.btn_ind    = {}
        self._rl_status = {}

        for i in range(1, N_RELAY + 1):
            cf = tk.Frame(ch_row, bg=CARD_BG, relief="raised", bd=2,
                          padx=10, pady=10)
            cf.pack(side="left", padx=5)

            tk.Label(cf, text=f"CH {i}", bg=CARD_BG, fg="#94a3b8",
                     font=("Segoe UI", 9, "bold")).pack()
            tk.Label(cf, text=f"RELAY {i}" if i != 4 else "RELAY BATT", bg=CARD_BG, fg="#1e293b",
                     font=("Segoe UI", 11, "bold")).pack(pady=(2, 8))

            b = self._flat_btn(cf, "OFF", RL_OFF_BG, font_size=10,
                                padx=8, pady=7, state="disabled", width=10,
                                raised=True,
                                command=lambda n=i: self._toggle_relay(n))
            b.config(activebackground=RL_OFF_BG)
            b.pack(fill="x", pady=(0, 6))
            self.btn_relay[i] = b

            srow = tk.Frame(cf, bg=CARD_BG)
            srow.pack(pady=(0, 8))
            d, doid = _make_dot(srow, 10, CARD_BG, "#94a3b8")
            d.pack(side="left", padx=(0, 5))
            l_st = tk.Label(srow, text="DANG TAT", bg=CARD_BG,
                             fg="#94a3b8", font=("Segoe UI", 9),
                             width=9, anchor="w")
            l_st.pack(side="left")
            self._rl_status[i] = (d, doid, l_st)

            tk.Frame(cf, bg=CARD_BD, height=1).pack(fill="x", pady=4)

            # LED: nút bấm nổi 3D (relief raised) -> click để test bật/tắt LED.
            lr = tk.Frame(cf, bg=CARD_BG, relief="raised", bd=2, cursor="hand2")
            lr.pack(pady=3, anchor="w", ipadx=4, ipady=2)
            cl = tk.Canvas(lr, width=16, height=16, bg=CARD_BG,
                            highlightthickness=0, cursor="hand2")
            ol = cl.create_oval(2, 2, 14, 14,
                                 fill=LED_OFF_FILL, outline=LED_OFF_LINE, width=1)
            cl.pack(side="left", padx=(2, 4))
            led_lbl = tk.Label(lr, text=f"LED {i}", bg=CARD_BG, fg="#475569",
                               font=("Segoe UI", 9, "bold"), cursor="hand2")
            led_lbl.pack(side="left", padx=(0, 2))
            for _w in (lr, cl, led_lbl):
                _w.bind("<ButtonPress-1>",
                        lambda e, f=lr: f.config(relief="sunken"))
                _w.bind("<ButtonRelease-1>",
                        lambda e, n=i, f=lr: (f.config(relief="raised"),
                                              self._toggle_led(n)))
            self.led_ind[i] = (cl, ol)

            # Trạng thái NÚT NHẤN của kênh này (gộp vào thẻ ĐIỀU KHIỂN RELAY)
            br = tk.Frame(cf, bg=CARD_BG)
            br.pack(pady=3, anchor="w")
            cb = tk.Canvas(br, width=14, height=14, bg=CARD_BG,
                            highlightthickness=0)
            ob = cb.create_rectangle(2, 2, 12, 12,
                                      fill=BTN_UP_FILL, outline=BTN_UP_LINE, width=1)
            cb.pack(side="left", padx=(0, 6))
            bt_lbl = tk.Label(br, text=f"BT {i}: Released", bg=CARD_BG,
                              fg="#16a34a", font=("Segoe UI", 9), anchor="w")
            bt_lbl.pack(side="left")
            self.btn_ind[i] = (cb, ob, bt_lbl)

        # ── QUICK ACTIONS ─────────────────────────────────────
        qa_body = self._card(top_row, "Thao Tác Nhanh",
                              side="left", fill="both", expand=True, padx=(0, 8))
        qa = tk.Frame(qa_body, bg=WHITE, padx=14, pady=12)
        qa.pack(fill="both", expand=True)

        self.btn_all_on = self._flat_btn(
            qa, "All On", "#22c55e", hover="#16a34a",
            font_size=10, padx=14, pady=10, state="disabled", raised=True,
            command=lambda: self._relay_all(True))
        self.btn_all_on.pack(fill="x", pady=3)

        self.btn_all_off = self._flat_btn(
            qa, "All Off", "#ef4444", hover="#dc2626",
            font_size=10, padx=14, pady=10, state="disabled", raised=True,
            command=lambda: self._relay_all(False))
        self.btn_all_off.pack(fill="x", pady=3)

        # Nút nổi 3D (relief raised) đồng bộ với các nút khác trên giao diện
        self.btn_rls = self._flat_btn(
            qa, "Sync Status Relay", "#64748b", hover="#475569",
            font_size=9, padx=8, pady=8, raised=True, state="disabled",
            command=self._push_relay_state)
        self.btn_rls.pack(fill="x", pady=3)

        # Nút bật/tắt tiếng bíp (test im lặng) - dùng được cả khi chưa kết nối
        self.btn_mute = self._flat_btn(
            qa, "Beep ON", BLUE_BTN, hover=BLUE_HOV,
            font_size=9, padx=8, pady=8, raised=True,
            command=self._toggle_mute)
        self.btn_mute.pack(fill="x", pady=3)

        # ── TEST CHỨC NĂNG CHUNG (I2C, Flash, DIP Switch) ─────
        fn_body = self._card(top_row, "Test Chức Năng Chung",
                             side="left", fill="both", expand=True)
        # Các nhãn kết quả đặt width cố định (theo số ký tự) -> nội dung thay
        # đổi không làm thẻ bị co giãn/nhảy layout.
        fn = tk.Frame(fn_body, bg=WHITE, padx=12, pady = 10)
        fn.pack(fill="both", expand=True)

        # --- I2C ---
        r1 = tk.Frame(fn, bg=WHITE)
        r1.pack(fill="x", pady=3)
        self.btn_test_i2c = self._flat_btn(
            r1, "Scan I2C", BLUE_BTN, hover=BLUE_HOV, font_size=9,
            padx=10, pady=5, width=9, raised=True, state="disabled",
            command=self._test_i2c)
        self.btn_test_i2c.pack(side="left")
        self.lbl_i2c = tk.Label(r1, text="ADDR: --", bg=WHITE, fg="#475569",
                                font=("Segoe UI", 9), width=18, anchor="w",
                                justify="left", wraplength=150)
        self.lbl_i2c.pack(side="left", padx=(8, 0))

        # --- FLASH ---
        r2 = tk.Frame(fn, bg=WHITE)
        r2.pack(fill="x", pady=3)
        self.btn_test_flash = self._flat_btn(
            r2, "Test Flash", BLUE_BTN, hover=BLUE_HOV, font_size=9,
            padx=10, pady=5, width=9, raised=True, state="disabled",
            command=self._test_flash)
        self.btn_test_flash.pack(side="left")
        self.lbl_flash = tk.Label(r2, text="--", bg=WHITE, fg="#475569",
                                  font=("Segoe UI", 9), width=18, anchor="w")
        self.lbl_flash.pack(side="left", padx=(8, 0))

        # --- DIP Switch (tự động cập nhật từ thiết bị, không cần nút test) ---
        tk.Frame(fn, bg=CARD_BD, height=1).pack(fill="x", pady=(6, 2))
        r3 = tk.Frame(fn, bg=WHITE)
        r3.pack(fill="x", pady=3)
        tk.Label(r3, text="Dip Switch:", bg=WHITE, fg="#475569",
                 font=("Segoe UI", 9, "bold")).pack(side="left")
        self.lbl_dip_addr = tk.Label(r3, text="ADDR: --", bg=WHITE, fg="#1e293b",
                                     font=("Segoe UI", 9, "bold"), width=20,
                                     anchor="w")
        self.lbl_dip_addr.pack(side="left", padx=(6, 0))

        # DIP 4 bit hiển thị trực tiếp (tự động cập nhật khi gạt switch)
        dipf = tk.Frame(fn, bg=WHITE)
        dipf.pack(fill="x", pady=(4, 0))
        self.dip_ind = {}
        for b in range(1, 5):
            cell = tk.Frame(dipf, bg=WHITE)
            cell.pack(side="left", padx=7)
            dot = tk.Canvas(cell, width=16, height=16, bg=WHITE,
                            highlightthickness=0)
            oid = dot.create_oval(2, 2, 14, 14, fill="#cbd5e1", outline="#94a3b8")
            dot.pack()
            tk.Label(cell, text=f"BIT{b}", bg=WHITE, fg="#6b7280",
                     font=("Segoe UI", 8)).pack()
            self.dip_ind[b] = (dot, oid)

        # ── TEST RS485 (thủ công: gửi text + loopback) ────────
        rs_body = self._card(top_row, "Test RS485",
                             side="left", fill="both", expand=True, padx=(8, 0))
        rs = tk.Frame(rs_body, bg=WHITE, padx=14, pady=12)
        rs.pack(fill="both", expand=True)
        # Đổi baud RS485 ngay trong thẻ (gửi "baud rs485 <n>" + cập nhật hiển thị)
        brow = tk.Frame(rs, bg=WHITE)
        brow.pack(fill="x")
        tk.Label(brow, text="Baud RS485:", bg=WHITE, fg="#6b7280",
                 font=("Segoe UI", 9)).pack(side="left")
        self.cbo_rs485_baud = ttk.Combobox(brow, width=8, state="readonly",
                                           values=[str(b) for b in BAUDRATES])
        self.cbo_rs485_baud.set("9600")
        self.cbo_rs485_baud.pack(side="left", padx=(4, 4))
        self.btn_rs485_baud = self._flat_btn(
            brow, "Đặt", BLUE_BTN, hover=BLUE_HOV, font_size=8,
            padx=10, pady=3, raised=True, state="disabled",
            command=self._set_rs485_baud)
        self.btn_rs485_baud.pack(side="left")
        # Baud RS485 hiện tại (hiển thị động)
        self.lbl_rs485_baud = tk.Label(rs, text="Baud RS485: 9600", bg=WHITE,
                                       fg="#e8833a", font=("Segoe UI", 8, "bold"))
        self.lbl_rs485_baud.pack(anchor="w", pady=(2, 6))
        tk.Label(rs, text="Text gửi qua RS485:", bg=WHITE, fg="#6b7280",
                 font=("Segoe UI", 9)).pack(anchor="w")
        self.ent_rs485 = tk.Entry(rs, font=("Segoe UI", 10), relief="solid", bd=1)
        self.ent_rs485.insert(0, RS485_TEST_TEXT)
        self.ent_rs485.pack(fill="x", pady=(2, 8))
        self.btn_rs485 = self._flat_btn(
            rs, "Gửi & Test Loopback", BLUE_BTN, hover=BLUE_HOV,
            font_size=10, padx=14, pady=8, state="disabled",
            raised=True, command=self._rs485_manual_test)
        self.btn_rs485.pack(fill="x")
        self.lbl_rs485 = tk.Label(rs, text="Chưa test", bg=WHITE, fg="#6b7280",
                                   font=("Segoe UI", 9), wraplength=200,
                                   justify="left")
        self.lbl_rs485.pack(anchor="w", pady=(8, 0))

        # ── BOTTOM ROW ────────────────────────────────────────
        bot_row = tk.Frame(content, bg=MAIN_BG)
        bot_row.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # KET QUA TEST
        tr_body = self._card(bot_row, "Kết Quả Test",
                              side="left", fill="both", expand=True, padx=(0, 8))
        tf = tk.Frame(tr_body, bg=WHITE, padx=6, pady=6)
        tf.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tf, columns=("res",), show="tree headings",
                                  selectmode="none", height=8)
        self.tree.heading("#0", text="Test")
        self.tree.heading("res", text="Kết quả")
        self.tree.column("#0", width=210)
        self.tree.column("res", width=80, anchor="center")
        self.tree.tag_configure("pass", foreground="#16a34a")
        self.tree.tag_configure("fail", foreground="#dc2626")
        self.tree.tag_configure("skip", foreground="#94a3b8")
        self.tree.tag_configure("run",  foreground="#2563eb")
        sb_t = ttk.Scrollbar(tf, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb_t.set)
        sb_t.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        self.lbl_sum = tk.Label(tr_body, text="Chưa chạy test",
                                 bg=WHITE, fg="#6b7280",
                                 font=("Segoe UI", 9), pady=6, padx=8, anchor="w")
        self.lbl_sum.pack(fill="x")

        # TERMINAL
        tm_outer = tk.Frame(bot_row, bg=WHITE,
                             highlightbackground=CARD_BD, highlightthickness=1)
        tm_outer.pack(side="left", fill="both", expand=True)

        tm_hdr = tk.Frame(tm_outer, bg=WHITE, padx=12, pady=8)
        tm_hdr.pack(fill="x")
        tk.Label(tm_hdr, text="Terminal",
                 bg=WHITE, fg=BLUE_ACC,
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Button(tm_hdr, text="Xóa Log",
                  bg=WHITE, fg="#6b7280",
                  activebackground="#f1f5f9", relief="solid", bd=1,
                  font=("Segoe UI", 8), padx=6, pady=1,
                  command=self._clear_log).pack(side="right")
        tk.Frame(tm_outer, bg=CARD_BD, height=1).pack(fill="x")

        # Input bar: pack TRƯỚC txt để không bị expand=True của txt đẩy ra ngoài
        inp = tk.Frame(tm_outer, bg="#1e293b", padx=8, pady=6)
        inp.pack(side="bottom", fill="x")
        tk.Label(inp, text=">", bg="#1e293b", fg="#22c55e",
                 font=("Consolas", 11, "bold")).pack(side="left", padx=(0, 6))
        self.ent_cmd = tk.Entry(inp, font=("Consolas", 10),
                                bg="#0f172a", fg="#e2e8f0",
                                insertbackground="white",
                                relief="flat", bd=0)
        self.ent_cmd.pack(side="left", fill="x", expand=True, ipady=4)
        self.ent_cmd.bind("<Return>", self._send_manual)
        self.ent_cmd.bind("<Up>",    self._cmd_hist_up)
        self.ent_cmd.bind("<Down>",  self._cmd_hist_down)
        self._flat_btn(inp, "Gửi", BLUE_BTN, hover=BLUE_HOV,
                        font_size=9, bold=True, padx=12, pady=4,
                        raised=True,
                        command=self._send_manual).pack(side="left", padx=(8, 0))

        # Quick commands — khớp với lệnh CLI có trên thiết bị (không có RTC/SHT45)
        qf = tk.Frame(tm_outer, bg="#0f172a", padx=6, pady=3)
        qf.pack(side="bottom", fill="x")
        tk.Frame(qf, bg="#1e293b", height=1).pack(fill="x", side="top", pady=(0, 3))
        tk.Label(qf, text="Quick:", bg="#0f172a", fg="#475569",
                 font=("Consolas", 8)).pack(side="left", padx=(0, 4))
        self._qcmd_btns = []
        _qcmds = [
            ("id",      "id"),       ("ver",  "ver"),   ("help", "help"), ("|", None),
            ("i2c",     "i2c"),      ("fwr",  "fwr"),   ("rls",  "rls"),  ("dip", "dip"), ("|", None),
            ("all ON",  "rl all on"),("all OFF","rl all off"), ("|", None),
            ("RS485\u21ba","rsl"),
        ]
        for _lbl, _cmd in _qcmds:
            if _cmd is None:
                tk.Label(qf, text="\u2502", bg="#0f172a", fg="#334155",
                         font=("Consolas", 9)).pack(side="left", padx=3)
            else:
                _b = tk.Button(qf, text=_lbl, bg="#1e3a5f", fg="#7dd3fc",
                               activebackground="#2563eb", activeforeground="white",
                               font=("Consolas", 8), relief="flat", bd=0,
                               padx=6, pady=2, state="disabled",
                               command=lambda c=_cmd: self._quick_send(c))
                _b.pack(side="left", padx=1, pady=1)
                self._qcmd_btns.append(_b)

        # Text log + scrollbar
        self.txt = tk.Text(tm_outer, bg="#0f172a", fg="#e2e8f0",
                            insertbackground="white", font=("Consolas", 9),
                            state="disabled", wrap="none", padx=8, pady=6,
                            width=40, height=8)
        self.txt.tag_configure("tx",  foreground="#60a5fa")
        self.txt.tag_configure("err", foreground="#f87171")
        sb_txt = ttk.Scrollbar(tm_outer, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb_txt.set)
        sb_txt.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True)

        # STATUS BAR
        sb_bar = tk.Frame(self, bg=HDR_BG, pady=3)
        sb_bar.pack(fill="x", side="bottom")
        tk.Frame(sb_bar, bg="#253445", height=1).pack(fill="x", side="top")
        self.lbl_status = tk.Label(sb_bar, text="* Chưa kết nối",
                                    bg=HDR_BG, fg="#64748b",
                                    font=("Segoe UI", 8), padx=12)
        self.lbl_status.pack(side="left")
        tk.Label(sb_bar, text="Smart PDU 2.0 Test App",
                 bg=HDR_BG, fg="#334155",
                 font=("Segoe UI", 8), padx=12).pack(side="right")

    # ----------------------------------------------------------
    #  Relay control
    # ----------------------------------------------------------
    def _set_relay_ctrl_state(self, enabled):
        st = "normal" if enabled else "disabled"
        for b in self.btn_relay.values():
            b.config(state=st)
        self.btn_all_on.config(state=st)
        self.btn_all_off.config(state=st)
        self.btn_rls.config(state=st)
        for _b in ("btn_rs485", "btn_rs485_baud", "btn_test_i2c", "btn_test_flash"):
            if hasattr(self, _b):
                getattr(self, _b).config(state=st)
        for _b in getattr(self, "_qcmd_btns", []):
            _b.config(state=st)

    def _rs485_manual_test(self):
        """Gửi text trong ô RS485 qua bus, kiểm tra có vòng về (loopback)."""
        if self.testing:
            return
        if not self.worker.is_open:
            messagebox.showwarning(APP_TITLE, "Chưa kết nối COM port")
            return
        text = self.ent_rs485.get().strip() or RS485_TEST_TEXT
        self.btn_rs485.config(state="disabled")
        self.lbl_rs485.config(text="Đang gửi...", fg="#2563eb")

        def _do():
            self.worker.clear_rx()
            resp = self.worker.send_cmd(f"rs485 {text}", 1.2)
            ok = text in resp

            def upd():
                self.btn_rs485.config(
                    state="normal" if (self.worker.is_open and not self.testing)
                    else "disabled")
                if ok:
                    self.lbl_rs485.config(text="OK - nhận đúng text vòng về",
                                          fg="#16a34a")
                else:
                    self.lbl_rs485.config(
                        text="FAIL - không nhận được\n(có thể loopback A-B / TX-RX)",
                        fg="#dc2626")
            self.after(0, upd)

        threading.Thread(target=_do, daemon=True).start()

    def _toggle_mute(self):
        """Bat/tat tiếng bip trên thiết bị (test im lặng)."""
        self._buzzer_muted = not self._buzzer_muted
        if self.worker.is_open:
            self.worker.write_line("beep off" if self._buzzer_muted else "beep on")
        self._update_mute_btn()

    def _update_mute_btn(self):
        if self._buzzer_muted:
            self.btn_mute.config(text="Beep OFF(Mute)", bg="#ef4444",
                                 activebackground="#dc2626")
        else:
            self.btn_mute.config(text="Beep ON", bg=BLUE_BTN,
                                 activebackground=BLUE_HOV)

    # ----------------------------------------------------------
    #  TEST CHUNG (I2C / Flash / Dip Switch)
    # ----------------------------------------------------------
    def _test_i2c(self):
        if self.testing or not self.worker.is_open:
            return
        self.btn_test_i2c.config(state="disabled")
        self.lbl_i2c.config(text="đang quét...", fg="#2563eb")

        def _do():
            resp = self.worker.send_cmd("i2c", 0.6)
            addrs = re.findall(r"I2C:\s*(0x[0-9A-Fa-f]+)", resp)

            def upd():
                self.btn_test_i2c.config(
                    state="normal" if (self.worker.is_open and not self.testing)
                    else "disabled")
                if addrs:
                    self.lbl_i2c.config(text="ADDR: " + ", ".join(addrs),
                                        fg="#16a34a")
                else:
                    self.lbl_i2c.config(text="không thấy thiết bị I2C",
                                        fg="#dc2626")
            self.after(0, upd)
        threading.Thread(target=_do, daemon=True).start()

    def _test_flash(self):
        if self.testing or not self.worker.is_open:
            return
        self.btn_test_flash.config(state="disabled")
        self.lbl_flash.config(text="đang test...", fg="#2563eb")

        def _do():
            resp = self.worker.send_cmd("fwr", 1.5)
            ok = "FLASH OK" in resp

            def upd():
                self.btn_test_flash.config(
                    state="normal" if (self.worker.is_open and not self.testing)
                    else "disabled")
                self.lbl_flash.config(
                    text="OK (ghi/đọc đạt)" if ok else "FAIL",
                    fg="#16a34a" if ok else "#dc2626")
            self.after(0, upd)
        threading.Thread(target=_do, daemon=True).start()

    def _update_relay_btn(self, n, on):
        self.relay_state[n] = on
        bg = RL_ON_BG if on else RL_OFF_BG
        self.btn_relay[n].config(text="ON" if on else "OFF",
                                  bg=bg, activebackground=bg)
        d, doid, lbl = self._rl_status[n]
        if on:
            d.itemconfig(doid, fill="#22c55e")
            lbl.config(text="ĐANG BẬT", fg="#16a34a")
        else:
            d.itemconfig(doid, fill="#94a3b8")
            lbl.config(text="ĐANG TẮT", fg="#94a3b8")
        for i in range(1, N_RELAY + 1):
            self._update_led_ind(i, self.relay_state[i])

    def _update_led_ind(self, n, on):
        self.led_state[n] = on
        c, oid = self.led_ind[n]
        c.itemconfig(oid,
                     fill=LED_ON_FILL if on else LED_OFF_FILL,
                     outline=LED_ON_LINE if on else LED_OFF_LINE)

    def _toggle_led(self, n):
        if self.testing or not self.worker.is_open:
            return
        target = not self.led_state[n]
        _cmd_on = not target if LED_ACTIVE_LOW else target
        self.worker.write_line(f"tled {n} {'on' if _cmd_on else 'off'}")
        self._update_led_ind(n, target)

    def _toggle_relay(self, n):
        if self.testing or not self.worker.is_open:
            return
        if time.time() < self._rl_busy_until:
            return
        self._rl_busy_until = time.time() + 0.6
        target = not self.relay_state[n]
        self.worker.write_line(f"rl {n} {'on' if target else 'off'}")
        # LED do firmware tu dieu khien bấm theo trang thai relay (led_update).
        self._update_relay_btn(n, target)

    def _push_relay_state(self):
        if self.testing or not self.worker.is_open:
            return
        if time.time() < self._rl_busy_until:
            return
        self._rl_busy_until = time.time() + N_RELAY * RL_SEQ_DELAY + 0.5
        self._log("[Sync trạng thái relay từ app xuống thiết bị...]\n", "tx")
        for i in range(1, N_RELAY + 1):
            self.worker.write_line(f"rl {i} {'on' if self.relay_state[i] else 'off'}")

    def _relay_all(self, on):
        """Điều khiển tuần tự relay 1->4, delay RL_SEQ_DELAY giữa mỗi relay.
        Tránh tăng dòng đột biến cho nguồn cấp relay (latching relay pulse 200ms/cái).
        Chạy trong thread riêng, nút ALL ON/OFF disable trong lúc chạy."""
        if self.testing or not self.worker.is_open:
            return
        if time.time() < self._rl_busy_until:
            return

        self.btn_all_on.config(state="disabled")
        self.btn_all_off.config(state="disabled")
        self._rl_busy_until = time.time() + N_RELAY * RL_SEQ_DELAY + 0.5

        def _seq():
            for i in range(1, N_RELAY + 1):
                if not self.worker.is_open:
                    break
                self.worker.write_line(f"rl {i} {'on' if on else 'off'}")
                self.after(0, lambda n=i: self._update_relay_btn(n, on))
                if i < N_RELAY:
                    time.sleep(RL_SEQ_DELAY)
            # Re-enable sau khi xong
            self.after(0, lambda: (
                self.btn_all_on.config(state="normal" if self.worker.is_open else "disabled"),
                self.btn_all_off.config(state="normal" if self.worker.is_open else "disabled"),
            ))

        threading.Thread(target=_seq, daemon=True).start()

    def _query_fw_ver(self):
        if not self.worker.is_open or self.testing:
            return
        def _do():
            resp = self.worker.send_cmd("ver", 0.4)
            m = re.search(r"FW\s+(\S+)", resp)
            if m:
                v = m.group(1)
                self.after(0, lambda: self._dev_info["Firmware:"].config(text=v))
        threading.Thread(target=_do, daemon=True).start()

    def _sync_state_from_device(self):
        """Khi kết nối: đọc trạng thái relay từ thiết bị (rls) và cập nhật
        hiển thị relay + LED trên app cho khớp (LED bấm theo trạng thái relay)."""
        if not self.worker.is_open or self.testing:
            return
        def _do():
            resp = self.worker.send_cmd("rls", 0.5)
            states = {(4 if m.group(1) == "_BATT" else int(m.group(1))): (m.group(2) == "ON")
                      for m in RX_RELAY_RE.finditer(resp)}
            dresp = self.worker.send_cmd("dip", 0.4)        # đọc DIP Switch hiện tại
            dm = RX_DIP_RE.search(dresp)
            dval = int(dm.group(1), 16) if dm else None

            def _apply():
                for n, s in states.items():
                    self._update_relay_btn(n, s)
                if dval is not None:
                    self._update_dip(dval)
                self._log("[Đã đồng bộ trạng thái relay/LED/Dip Switch từ thiết bị]\n", "tx")
            if states or dval is not None:
                self.after(0, _apply)
        threading.Thread(target=_do, daemon=True).start()

    # ----------------------------------------------------------
    #  Button indicators
    # ----------------------------------------------------------
    def _update_btn_ind(self, n, down):
        c, oid, lbl = self.btn_ind[n]
        c.itemconfig(oid,
                     fill=BTN_DOWN_FILL if down else BTN_UP_FILL,
                     outline=BTN_DOWN_LINE if down else BTN_UP_LINE)
        if down:
            lbl.config(text=f"BT {n}: Pressed", fg="#3b82f6")
        else:
            lbl.config(text=f"BT {n}: Released", fg="#16a34a")

    def _set_btn_error(self, n):
        c, oid, lbl = self.btn_ind[n]
        c.itemconfig(oid, fill="#ef4444", outline="#dc2626")
        lbl.config(text=f"BT {n}: Lỗi", fg="#ef4444")

    def _on_firmware_reset(self):
        self._btn_down_time.clear()
        self._btn_pre_relay.clear()
        for i in range(1, N_RELAY + 1):
            self._update_relay_btn(i, False)
        self._log("[Firmware đã reset - trạng thái relay về OFF]\n", "err")

    def _scan_relay_rx(self, text):
        if RX_FW_RESET_RE.search(text):
            self._on_firmware_reset()
            return

        # Relay do FIRMWARE điều khiển khi giữ/nhả nút (giữ [1s,3s) mới đảo;
        # giữ quá lâu/nhiều nút = lỗi, KHÔNG đảo). App chỉ hiển thị trạng thái
        # nút và cập nhật relay qua thông báo "RLn:" từ firmware -> không tự
        # gửi lệnh relay ở đây để tránh điều khiển trùng.
        for m in RX_BTN_RE.finditer(text):
            n    = int(m.group(1))
            down = (m.group(2) == "DOWN")
            self._update_btn_ind(n, down)
            if down and self._bt_test_active:
                self._bt_test_pressed.add(n)   # ghi nhận cho bước test nút

        for m in RX_BTN_ERR_HOLD.finditer(text):
            n = int(m.group(1))
            self._btn_down_time.pop(n, None)
            self._btn_pre_relay.pop(n, None)
            self._set_btn_error(n)
            self._log(f"[BT{n}: Giữ quá lâu - không điều khiển relay]\n", "err")

        if RX_BTN_ERR_MULTI.search(text):
            self._btn_down_time.clear()
            self._btn_pre_relay.clear()
            for i in range(1, N_RELAY + 1):
                self._set_btn_error(i)
            self._log("[BT_ERR: Nhấn nhiều nút - không điều khiển relay]\n", "err")

        for m in RX_RELAY_RE.finditer(text):
            ch_key = m.group(1)
            ch = 4 if ch_key == "_BATT" else int(ch_key)
            self._update_relay_btn(ch, m.group(2) == "ON")
        for m in RX_LED_RE.finditer(text):
            _gpio_on = m.group(2) == "ON"
            self._update_led_ind(int(m.group(1)), not _gpio_on if LED_ACTIVE_LOW else _gpio_on)
        for m in RX_DIP_RE.finditer(text):
            val = int(m.group(1), 16)
            # Trong bước test DIP: ghi nhận bit nào vừa THAY ĐỔI (gạt)
            if self._dip_test_active and self._dip_last_value is not None:
                changed = self._dip_last_value ^ val
                for b in range(1, 5):
                    if changed & (1 << (b - 1)):
                        self._dip_test_seen.add(b)
            self._dip_last_value = val
            self._update_dip(val)
        for m in RX_RS485_BAUD_RE.finditer(text):
            self._set_rs485_baud_display(m.group(1))

    def _set_rs485_baud_display(self, baud):
        """Cập nhật baud RS485 ở thẻ Test RS485 (nhãn + ô chọn)."""
        baud = str(baud)
        self.lbl_rs485_baud.config(text=f"Baud RS485: {baud} ")
        if hasattr(self, "cbo_rs485_baud") and baud in self.cbo_rs485_baud["values"]:
            self.cbo_rs485_baud.set(baud)

    def _set_rs485_baud(self):
        """Đổi baud RS485 trên thiết bị từ ô chọn trong thẻ Test RS485."""
        if self.testing or not self.worker.is_open:
            return
        baud = self.cbo_rs485_baud.get()
        self.worker.write_line(f"baud rs485 {baud}")
        self._set_rs485_baud_display(baud)
        self._log(f"[Đổi baud RS485 -> {baud}]\n", "tx")

    def _update_dip(self, value):
        """Cập nhật hiển thị DIP Switch (4 bit + địa chỉ) khi gạt switch."""
        bits = []
        for b in range(1, 5):
            on = bool(value & (1 << (b - 1)))    # bit0 = BIT1
            dot, oid = self.dip_ind[b]
            dot.itemconfig(oid, fill="#22c55e" if on else "#cbd5e1",
                           outline="#15803d" if on else "#94a3b8")
            bits.append("1" if on else "0")
        # bits hiển thị BIT1..BIT4 trái->phải; địa chỉ = giá trị 4-bit
        self.lbl_dip_addr.config(
            text=f"ADDR: 0x{value:X} ({value}) b{''.join(bits)}")

    # ----------------------------------------------------------
    #  COM port
    # ----------------------------------------------------------
    def _refresh_ports(self):
        ports = serial.tools.list_ports.comports()
        items = [f"{p.device} - {p.description}" for p in ports]
        self.cbo_port["values"] = items
        if items and not self.cbo_port.get():
            self.cbo_port.current(0)

    def _sel_port(self):
        v = self.cbo_port.get()
        return v.split(" - ")[0] if v else ""

    def _toggle_conn(self):
        if self.worker.is_open:
            self.worker.close()
            self._rx_scan_buf = ""
            self.btn_conn.config(text="Kết Nối", bg=BLUE_BTN,
                                  activebackground=BLUE_HOV)
            self.btn_run.config(state="disabled")
            self._set_relay_ctrl_state(False)
            self._conn_dot.itemconfig(self._conn_dot_oid, fill="#64748b")
            self._lbl_conn.config(text="Chưa kết nối", fg="#94a3b8")
            self.lbl_status.config(text="* Chưa kết nối", fg="#64748b")
            self._dev_info["Firmware:"].config(text="--")
            return

        if self._connecting:
            # Nút đang đóng vai trò "Hủy" trong lúc mở cổng/xác thực
            self._cancel_connect()
            return

        port = self._sel_port()
        if not port:
            messagebox.showwarning(APP_TITLE, "Chưa chọn COM port")
            return
        baud = int(self.cbo_baud.get() or BAUD)

        # Mở cổng CHẠY NỀN (thread riêng): một số cổng COM ảo (vd Bluetooth
        # SPP) có thể treo nhiều giây/vô thời hạn ngay tại bước mở cổng ->
        # nếu gọi serial.Serial() trực tiếp trên main thread, cả app sẽ bị
        # đứng (không bấm được gì, kể cả nút Hủy). Mở nền + nút Hủy cho phép
        # người dùng bỏ attempt bị treo và chọn cổng COM khác ngay.
        self._connecting = True
        self._connect_cancelled = False
        self.cbo_port.config(state="disabled")
        self.cbo_baud.config(state="disabled")
        self.btn_refresh_port.config(state="disabled")
        self.btn_conn.config(text="Hủy Kết Nối", bg="#f59e0b",
                              activebackground="#d97706", state="normal",
                              command=self._cancel_connect)
        self._lbl_conn.config(text="Đang mở cổng...", fg="#f59e0b")
        self._conn_dot.itemconfig(self._conn_dot_oid, fill="#f59e0b")
        self.lbl_status.config(text=f"* Đang mở {port} @ {baud}...",
                               fg="#f59e0b")
        self._log(f"[Đang mở {port} @ {baud}...]\n", "tx")

        def _do_open():
            try:
                ser = serial.Serial(port, baud, timeout=0.1)
            except Exception as e:
                self.after(0, lambda: self._on_open_result(None, e, port, baud))
                return
            self.after(0, lambda: self._on_open_result(ser, None, port, baud))

        threading.Thread(target=_do_open, daemon=True).start()
        # Watchdog: mở cổng quá lâu (thường gặp ở cổng Bluetooth/ảo bị treo)
        # -> cảnh báo người dùng có thể bấm Hủy để chọn cổng khác.
        self._connect_timeout_id = self.after(4000, self._on_connect_slow)

    def _on_connect_slow(self):
        self._connect_timeout_id = None
        if self._connecting and not self._connect_cancelled:
            self.lbl_status.config(
                text=("* Mở cổng quá lâu - có thể đây là cổng COM ảo "
                      "(Bluetooth/...) bị treo. Bấm 'Hủy Kết Nối' để chọn "
                      "cổng khác."), fg="#ef4444")
            self._log("[Cảnh báo: mở cổng quá lâu - bấm Hủy Kết Nối để chọn "
                       "cổng COM khác]\n", "err")

    def _cancel_connect(self):
        """Người dùng bấm Hủy trong lúc đang mở cổng / xác thực ID.
        Lưu ý: nếu serial.Serial() đang treo trong thread nền, KHÔNG có cách
        nào dừng cuộc gọi đó giữa chừng (Python không kill thread blocking
        I/O) -> ta chỉ bỏ qua kết quả của attempt đó (cờ _connect_cancelled)
        và trả UI về trạng thái sẵn sàng để người dùng thử cổng COM khác
        ngay; thread cũ (daemon) tự kết thúc âm thầm khi nó trả về."""
        self._connect_cancelled = True
        if self._connect_timeout_id is not None:
            self.after_cancel(self._connect_timeout_id)
            self._connect_timeout_id = None
        try:
            self.worker.close()
        except Exception:
            pass
        self._connecting = False
        self._rx_scan_buf = ""
        self.cbo_port.config(state="readonly")
        self.cbo_baud.config(state="readonly")
        self.btn_refresh_port.config(state="normal")
        self.btn_conn.config(text="Kết Nối", bg=BLUE_BTN,
                              activebackground=BLUE_HOV, state="normal",
                              command=self._toggle_conn)
        self.btn_run.config(state="disabled")
        self._set_relay_ctrl_state(False)
        self._conn_dot.itemconfig(self._conn_dot_oid, fill="#64748b")
        self._lbl_conn.config(text="Chưa kết nối", fg="#94a3b8")
        self.lbl_status.config(text="* Đã hủy kết nối - chọn cổng COM khác",
                               fg="#94a3b8")
        self._log("[Đã hủy kết nối - có thể chọn cổng COM khác]\n", "err")

    def _on_open_result(self, ser, err, port, baud):
        if self._connect_timeout_id is not None:
            self.after_cancel(self._connect_timeout_id)
            self._connect_timeout_id = None

        if self._connect_cancelled:
            # Người dùng đã Hủy trước khi mở xong (UI đã được reset rồi) ->
            # nếu cổng cuối cùng vẫn mở được thì đóng luôn, không dùng nữa.
            if ser is not None:
                try:
                    ser.close()
                except Exception:
                    pass
            return

        if err is not None:
            self._connecting = False
            self.cbo_port.config(state="readonly")
            self.cbo_baud.config(state="readonly")
            self.btn_refresh_port.config(state="normal")
            self.btn_conn.config(text="Kết Nối", bg=BLUE_BTN,
                                  activebackground=BLUE_HOV,
                                  command=self._toggle_conn)
            self._conn_dot.itemconfig(self._conn_dot_oid, fill="#ef4444")
            self._lbl_conn.config(text="Chưa kết nối", fg="#94a3b8")
            self.lbl_status.config(text="* Chưa kết nối", fg="#64748b")
            self._log(f"[Không mở được {port}: {err}]\n", "err")
            messagebox.showerror(APP_TITLE, f"Không mở được {port}:\n{err}")
            return

        # Mở cổng OK -> gắn vào worker, bắt đầu thread đọc, rồi XÁC THỰC ID
        # trước khi cho điều khiển (tránh nhầm cổng COM).
        self.worker.ser = ser
        self.worker._stop.clear()
        self.worker._thread = threading.Thread(
            target=self.worker._reader, daemon=True)
        self.worker._thread.start()

        self._lbl_conn.config(text="Đang kiểm tra...", fg="#f59e0b")
        self.lbl_status.config(text=f"* {port}: đang xác thực thiết bị...",
                               fg="#f59e0b")
        self._log(f"[Mở {port} @ {baud} - kiểm tra ID thiết bị...]\n", "tx")

        def _verify():
            resp = self.worker.send_cmd("id", 0.6)
            ok = DEVICE_ID_SIG in resp
            self.after(0, lambda: self._on_verify_result(ok, port, baud))
        threading.Thread(target=_verify, daemon=True).start()

    def _on_verify_result(self, ok, port, baud):
        if self._connect_cancelled:
            return    # đã Hủy, UI đã được reset trong _cancel_connect
        self._connecting = False
        self.cbo_port.config(state="readonly")
        self.cbo_baud.config(state="readonly")
        self.btn_refresh_port.config(state="normal")
        self.btn_conn.config(state="normal")
        if not self.worker.is_open:
            return    # người dùng đã ngắt trong lúc kiểm tra
        if not ok:
            # SAI thiết bi / nhầm cổng COM -> đóng cổng, cảnh báo
            self.worker.close()
            self.btn_conn.config(text="Kết Nối", bg=BLUE_BTN,
                                  activebackground=BLUE_HOV,
                                  command=self._toggle_conn)
            self._conn_dot.itemconfig(self._conn_dot_oid, fill="#ef4444")
            self._lbl_conn.config(text="SAI thiết bị/baud!", fg="#ef4444")
            self.lbl_status.config(
                text=f"* {port} @ {baud}: không nhận đúng Smart PDU 2.0",
                fg="#ef4444")
            self._log(f"[{port} @ {baud} không phản hồi đúng ID - đã ngắt]\n",
                      "err")
            messagebox.showwarning(
                APP_TITLE,
                f"Cổng {port} @ {baud} KHÔNG nhận đúng thiết bị Smart PDU 2.0!\n\n"
                "Nguyên nhân có thể:\n"
                f"  - Chọn nhầm cổng COM (không phải thiết bị này)\n"
                f"  - SAI baud rate: thiết bị dùng Console 115200\n\n"
                "Hãy kiểm tra lại cổng COM và đặt baud = 115200.")
            return

        # ĐÚNG thiết bị -> hoàn tất kết nối
        self.btn_conn.config(text="Ngắt Kết Nối", bg="#ef4444",
                              activebackground="#dc2626",
                              command=self._toggle_conn)
        self.btn_run.config(state="normal")
        self._set_relay_ctrl_state(True)
        self._conn_dot.itemconfig(self._conn_dot_oid, fill="#22c55e")
        self._lbl_conn.config(text="✓ Smart PDU 2.0", fg="#22c55e")
        self.lbl_status.config(
            text=f"* Đã kết nối thiết bị Smart PDU 2.0  |  {port} @ {baud}",
            fg="#16a34a")
        self._dev_info["Baud:"].config(text=str(baud))
        self._log(f"[Xác thực OK - Smart PDU 2.0 trên {port} @ {baud}]\n", "tx")
        # Đồng bộ trạng thái tắt/bật tiếng bip xuống thiết bị
        self.worker.write_line("beep off" if self._buzzer_muted else "beep on")
        # Truy vấn baud RS485 hiện tại để hiển thị động
        self.worker.write_line("baud")
        # Đọc firmware trước (lúc thiết bị còn rảnh), sau đó mới sync relay.
        self._after_fwver = self.after(300, self._query_fw_ver)
        self._after_sync = self.after(1300, self._sync_state_from_device)

    # ----------------------------------------------------------
    #  Terminal
    # ----------------------------------------------------------
    def _send_manual(self, _evt=None):
        if self.testing:
            return
        cmd = self.ent_cmd.get().strip()
        if not cmd:
            return
        if not self.worker.is_open:
            messagebox.showwarning(APP_TITLE, "Chưa kết nối COM port")
            return
        self.worker.write_line(cmd)
        if not self._cmd_history or self._cmd_history[-1] != cmd:
            self._cmd_history.append(cmd)
            if len(self._cmd_history) > 50:
                self._cmd_history.pop(0)
        self._cmd_hist_idx = -1
        # Nếu người dùng đổi baud RS485 qua terminal -> đồng bộ nhãn hiển thị ngay
        # (range khớp firmware 1200..921600; firmware cũng echo "RS485 BAUD:" lại)
        mb = re.match(r"\s*baud\s+rs485\s+(\d+)\s*$", cmd, re.IGNORECASE)
        if mb and 1200 <= int(mb.group(1)) <= 921600:
            self._set_rs485_baud_display(mb.group(1))
        self.ent_cmd.delete(0, "end")

    def _cmd_hist_up(self, _evt=None):
        """Phím ↑: duyệt lịch sử lệnh về phía cũ hơn."""
        if not self._cmd_history:
            return "break"
        if self._cmd_hist_idx < len(self._cmd_history) - 1:
            self._cmd_hist_idx += 1
        self.ent_cmd.delete(0, "end")
        self.ent_cmd.insert(0, self._cmd_history[-(self._cmd_hist_idx + 1)])
        return "break"

    def _cmd_hist_down(self, _evt=None):
        """Phím ↓: duyệt lịch sử lệnh về phía mới hơn."""
        if self._cmd_hist_idx > 0:
            self._cmd_hist_idx -= 1
            self.ent_cmd.delete(0, "end")
            self.ent_cmd.insert(0, self._cmd_history[-(self._cmd_hist_idx + 1)])
        else:
            self._cmd_hist_idx = -1
            self.ent_cmd.delete(0, "end")
        return "break"

    def _quick_send(self, cmd):
        """Gửi lệnh từ nút quick command trong terminal."""
        if not self.worker.is_open or self.testing:
            return
        self.worker.write_line(cmd)
        if not self._cmd_history or self._cmd_history[-1] != cmd:
            self._cmd_history.append(cmd)
            if len(self._cmd_history) > 50:
                self._cmd_history.pop(0)
        self._cmd_hist_idx = -1

    def _clear_log(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.config(state="disabled")

    def _log(self, text, tag=""):
        self.txt.config(state="normal")
        self.txt.insert("end", text, tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def _poll_log(self):
        try:
            while True:
                tag, text = self.log_queue.get_nowait()
                self._log(text, tag)
                if tag == "rx":
                    self._rx_scan_buf += text
        except queue.Empty:
            pass

        if self._rx_scan_buf:
            lines = self._rx_scan_buf.split('\n')
            self._rx_scan_buf = lines[-1]
            for line in lines[:-1]:
                self._scan_relay_rx(line)

        self.after(50, self._poll_log)

    # ----------------------------------------------------------
    #  Test runner
    # ----------------------------------------------------------
    def _run_tests(self):
        if self.testing or not self.worker.is_open:
            return
        serial_no = self.ent_serial.get().strip()
        if not serial_no:
            messagebox.showwarning(APP_TITLE,
                                   "Nhập Serial/Mã thiết bị trước khi RUN TEST")
            self.ent_serial.focus_set()
            return
        self._serial_no = serial_no
        # Hủy các tác vụ ngầm đặt lịch lúc kết nối (đọc FW / sync relay) nếu
        # chưa chạy -> tránh đụng độ lệnh 'ver' của test ngay sau khi kết nối.
        for _aid in ("_after_fwver", "_after_sync"):
            if getattr(self, _aid, None) is not None:
                try:
                    self.after_cancel(getattr(self, _aid))
                except Exception:
                    pass
                setattr(self, _aid, None)
        self.testing = True
        self._stop_test.clear()
        self.btn_run.config(text="Stop", bg="#ef4444",
                             activebackground="#dc2626",
                             command=self._stop_tests)
        self.btn_conn.config(state="disabled")
        self.ent_cmd.config(state="disabled")
        self._set_relay_ctrl_state(False)

        self.tree.delete(*self.tree.get_children())
        for i, t in enumerate(TESTS):
            self.tree.insert("", "end", iid=str(i), text=t["name"], values=("...",))

        self.lbl_sum.config(text="Đang chạy test... (bấm STOP để dừng)",
                             fg="#2563eb")
        threading.Thread(target=self._test_thread, daemon=True).start()

    def _stop_tests(self):
        self._stop_test.set()
        self.btn_run.config(state="disabled")
        self.lbl_sum.config(text="Đang dừng test...", fg="#f59e0b")

    def _record(self, i, token, results, cell=None):
        """Ghi kết quả 1 bước. token: 'pass'|'fail'|'skip'|'stop'.
        cell = chuỗi ghi vào Excel (mặc định theo token, vd skip -> 'Bỏ qua').
        results = list các tuple (token, cell)."""
        tree_txt = {"pass": "PASS", "fail": "FAIL",
                    "skip": "BO QUA", "stop": "-"}.get(token, "-")
        tag = {"pass": "pass", "fail": "fail",
               "skip": "skip", "stop": "skip"}.get(token, "skip")
        if cell is None:
            cell = {"pass": "PASS", "fail": "FAIL",
                    "skip": "Bo qua", "stop": "-"}.get(token, "-")
        self._set_row(i, tree_txt, tag)
        results.append((token, cell))

    def _test_thread(self):
        stopped = False
        results = []
        fw_ver = ""
        # Chờ thiết bị rảnh (vd vừa sync relay khi kết nối - mỗi relay pulse
        # 200ms) và xả buffer trước khi chạy, tránh 'ver' (test đầu) bị kẹt
        # phía sau hàng lệnh relay -> fail oan.
        while time.time() < self._rl_busy_until and not self._stop_test.is_set():
            time.sleep(0.1)
        self.worker.clear_rx()

        # Đọc firmware 1 lần để ghi vào báo cáo (KHÔNG phải 1 bước test).
        if not self._stop_test.is_set():
            _resp = self.worker.send_cmd("ver", 0.5)
            _m = re.search(r"FW\s+(\S+)", _resp)
            if _m:
                fw_ver = _m.group(1)
                self.after(0, lambda v=fw_ver:
                           self._dev_info["Firmware:"].config(text=v))

        for i, step in enumerate(TESTS):
            if self._stop_test.is_set():
                stopped = True
                self._record(i, "stop", results)
                continue

            self._set_row(i, "RUN", "run")
            kind = step["kind"]

            if kind == "auto":
                resp = self.worker.send_cmd(step["cmd"], step["wait"])
                ok = (all(s in resp for s in step["must"])
                      and not any(s in resp for s in step["must_not"]))
                self._record(i, "pass" if ok else "fail", results)

            elif kind == "led":
                token, cell = self._run_led_test()
                stopped = stopped or (token == "stop")
                self._record(i, token, results, cell)

            elif kind == "button":
                token, cell = self._run_button_test()
                stopped = stopped or (token == "stop")
                self._record(i, token, results, cell)

            elif kind == "dip":
                token, cell = self._run_dip_test()
                stopped = stopped or (token == "stop")
                self._record(i, token, results, cell)

            elif kind == "rs485":
                token, cell = self._run_rs485_test(step["text"])
                stopped = stopped or (token == "stop")
                self._record(i, token, results, cell)

        tokens = [t for t, _ in results]
        cells = [c for _, c in results]
        n_pass = tokens.count("pass")
        n_fail = tokens.count("fail")
        n_skip = tokens.count("skip")
        n_total = n_pass + n_fail + n_skip

        def done():
            self.testing = False
            self.btn_run.config(text="Run Test", bg=BLUE_BTN,
                                 activebackground=BLUE_HOV,
                                 command=self._run_tests, state="normal")
            self.btn_conn.config(state="normal")
            self.ent_cmd.config(state="normal")
            self._set_relay_ctrl_state(self.worker.is_open)
            summ = (f"[{self._serial_no}] {n_pass} PASS / {n_fail} FAIL "
                    f"/ {n_skip} SKIP")
            if stopped:
                ket = "ĐÃ DỪNG GIỮA CHỪNG"
                self.lbl_sum.config(text=summ + "  --  ĐÃ DỪNG", fg="#f59e0b")
            else:
                board_ok = (n_fail == 0)
                ket = "BOARD OK" if board_ok else "CÓ LỖI"
                self.lbl_sum.config(
                    text=summ + ("  --  BOARD OK" if board_ok else "  --  CÓ LỖI"),
                    fg="#16a34a" if board_ok else "#dc2626")
            # Thông báo hoàn tất + chọn lưu Excel hay bỏ qua
            if messagebox.askyesno(
                    APP_TITLE,
                    "Test hoàn tất!\n\n"
                    f"Serial: {self._serial_no}\n"
                    f"Kết quả: {n_pass} PASS / {n_fail} FAIL / {n_skip} SKIP\n"
                    f"Kết luận: {ket}\n\n"
                    "Lưu kết quả vào file Excel?"):
                self._save_report(fw_ver, cells, n_pass, n_fail, n_total, stopped)
            else:
                self._log("[Đã bỏ qua lưu report (người dùng chọn)]\n", "tx")
            self.ent_serial.delete(0, "end")
            self.ent_serial.focus_set()
        self.after(0, done)

    # ----------------------------------------------------------
    #  Interactive test helpers
    #  (Dialog cập nhật trên MAIN thread: test thread chỉ gọi
    #   self.after(0, show) MỘT lần rồi ev.wait(). Không gọi tkinter
    #   liên tục từ thread phụ -> tránh treo GUI.)
    # ----------------------------------------------------------
    def _dialog_place(self, win):
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 3
        win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        try:
            win.transient(self)
            win.attributes("-topmost", True)
            win.lift()
            win.focus_force()
        except Exception:
            pass

    def _ask_user(self, title, message, buttons):
        """Hoi nguoi dung. Tra ve value cua nut da chon, 'stop' neu dung test."""
        res = {"v": None}
        ev = threading.Event()

        def show():
            win = tk.Toplevel(self)
            win.title(title)
            win.configure(bg=WHITE)
            win.resizable(False, False)
            tk.Label(win, text=message, bg=WHITE, fg="#1e293b",
                     font=("Segoe UI", 11), justify="left",
                     wraplength=420, padx=24, pady=18).pack()
            bf = tk.Frame(win, bg=WHITE)
            bf.pack(pady=(0, 16))

            def pick(v):
                if not ev.is_set():
                    res["v"] = v
                    ev.set()
                try:
                    win.destroy()
                except Exception:
                    pass

            for (lbl, val, bg) in buttons:
                tk.Button(bf, text=lbl, bg=bg, fg="white", activebackground=bg,
                          font=("Segoe UI", 10, "bold"), relief="flat", bd=0,
                          padx=18, pady=8,
                          command=lambda vv=val: pick(vv)).pack(side="left", padx=8)
            win.protocol("WM_DELETE_WINDOW", lambda: pick("skip"))
            self._dialog_place(win)

            def tick():
                if ev.is_set():
                    return
                if self._stop_test.is_set():
                    pick("stop")
                    return
                win.after(150, tick)
            tick()

        self.after(0, show)
        ev.wait()
        return res["v"]

    def _run_led_test(self):
        # B1: thông báo để người test CHUẨN BỊ quan sát trước khi LED sáng
        ans = self._ask_user(
            "Kiểm tra LED",
            "Chuẩn bị quan sát 4 đèn LED trên board.\n"
            "Khi sẵn sàng, bấm 'Bắt đầu' -> LED 1 -> 4 sẽ sáng lần lượt,\n"
            "sau đó xác nhận đèn nào không sáng.",
            [("Bắt đầu", "go", "#16a34a"), ("Bỏ qua", "skip", "#64748b")])
        if ans == "stop":
            return ("stop", None)
        if ans != "go":
            return ("skip", "Bo qua")

        # B2: bật LED 1->4 lần lượt trên board (lệnh trực tiếp tled)
        self.worker.clear_rx()
        # LED active-low: lệnh làm LED sáng là "tled off" (pin LOW).
        _on  = "off" if LED_ACTIVE_LOW else "on"
        _off = "on"  if LED_ACTIVE_LOW else "off"
        for n in range(1, N_RELAY + 1):
            if self._stop_test.is_set():
                return ("stop", None)
            if self.worker.is_open:
                self.worker.write_line(f"tled {n} {_on}")
            time.sleep(0.5)

        res = {"token": None, "cell": None}
        ev = threading.Event()

        def show():
            fail = {n: False for n in range(1, N_RELAY + 1)}
            win = tk.Toplevel(self)
            win.title("Kiểm tra LED")
            win.configure(bg=WHITE)
            win.resizable(False, False)
            tk.Label(win, text="LED 1 -> 4 đã bật lần lượt trên board.\n"
                               "Click vào LED nào KHÔNG sáng để đánh dấu lỗi.",
                     bg=WHITE, fg="#1e293b", font=("Segoe UI", 11),
                     justify="left", padx=24).pack(pady=(16, 10))
            cellf = tk.Frame(win, bg=WHITE)
            cellf.pack(pady=4)
            cells = {}

            def toggle(n):
                fail[n] = not fail[n]
                dot, oid, lbl = cells[n]
                dot.itemconfig(oid, fill="#ef4444" if fail[n] else "#22c55e")
                lbl.config(text=f"LED{n}\n" + ("LOI" if fail[n] else "OK"),
                           fg="#ef4444" if fail[n] else "#16a34a")

            for n in range(1, N_RELAY + 1):
                cell = tk.Frame(cellf, bg=WHITE, cursor="hand2")
                cell.pack(side="left", padx=10)
                dot = tk.Canvas(cell, width=22, height=22, bg=WHITE,
                                highlightthickness=0)
                oid = dot.create_oval(2, 2, 20, 20, fill="#22c55e", outline="")
                dot.pack()
                lbl = tk.Label(cell, text=f"LED{n}\nOK", bg=WHITE, fg="#16a34a",
                               font=("Segoe UI", 9), justify="center")
                lbl.pack()
                cells[n] = (dot, oid, lbl)
                for w in (cell, dot, lbl):
                    w.bind("<Button-1>", lambda e, nn=n: toggle(nn))

            bf = tk.Frame(win, bg=WHITE)
            bf.pack(pady=(10, 16))

            def finish(kind):
                if not ev.is_set():
                    if kind == "confirm":
                        bad = [n for n in range(1, N_RELAY + 1) if fail[n]]
                        if bad:
                            res["token"] = "fail"
                            res["cell"] = "Loi: " + ",".join(f"LED{n}" for n in bad)
                        else:
                            res["token"] = "pass"
                            res["cell"] = "PASS"
                    elif kind == "skip":
                        res["token"] = "skip"
                        res["cell"] = "Bo qua"
                    else:  # stop
                        res["token"] = "stop"
                        res["cell"] = None
                    ev.set()
                try:
                    win.destroy()
                except Exception:
                    pass

            tk.Button(bf, text="Xác nhận", bg="#16a34a", fg="white",
                      activebackground="#16a34a", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("confirm")).pack(side="left", padx=8)
            tk.Button(bf, text="Bỏ qua", bg="#64748b", fg="white",
                      activebackground="#64748b", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("skip")).pack(side="left", padx=8)
            win.protocol("WM_DELETE_WINDOW", lambda: finish("skip"))
            self._dialog_place(win)

            def tick():
                if ev.is_set():
                    return
                if self._stop_test.is_set():
                    finish("stop")
                    return
                win.after(200, tick)
            tick()

        self.after(0, show)
        ev.wait()
        for n in range(1, N_RELAY + 1):     # tắt hết LED sau khi kiểm tra
            if self.worker.is_open:
                self.worker.write_line(f"tled {n} {_off}")
        return (res["token"], res["cell"])

    def _run_button_test(self):
        self._bt_test_pressed = set()
        self._bt_test_active = True
        res = {"token": None, "cell": None}
        ev = threading.Event()

        def show():
            manual = set()
            win = tk.Toplevel(self)
            win.title("Kiểm tra nút nhấn")
            win.configure(bg=WHITE)
            win.resizable(False, False)
            tk.Label(win, text="Nhấn lần lượt các nút BT1 -> BT4 trên thiết bị\n"
                               "(chạm chuyển xanh). Đủ 4 nút sẽ tự động PASS.\n"
                               "Hoặc click vào ô để đánh dấu thủ công.",
                     bg=WHITE, fg="#1e293b", font=("Segoe UI", 11),
                     justify="left", padx=24).pack(pady=(16, 10))
            statf = tk.Frame(win, bg=WHITE)
            statf.pack(pady=4)
            cells = {}

            def manual_mark(n):
                manual.add(n)

            for n in range(1, N_RELAY + 1):
                cell = tk.Frame(statf, bg=WHITE, cursor="hand2")
                cell.pack(side="left", padx=12)
                dot = tk.Canvas(cell, width=20, height=20, bg=WHITE,
                                highlightthickness=0)
                oid = dot.create_oval(2, 2, 18, 18, fill="#cbd5e1", outline="")
                dot.pack()
                lbl = tk.Label(cell, text=f"BT{n}", bg=WHITE,
                               font=("Segoe UI", 9))
                lbl.pack()
                cells[n] = (dot, oid)
                for w in (cell, dot, lbl):
                    w.bind("<Button-1>", lambda e, nn=n: manual_mark(nn))

            bf = tk.Frame(win, bg=WHITE)
            bf.pack(pady=(10, 16))

            def finish(kind):
                if not ev.is_set():
                    marked = set(self._bt_test_pressed) | manual
                    if kind in ("confirm", "auto"):
                        miss = [n for n in range(1, N_RELAY + 1) if n not in marked]
                        if kind == "auto" or not miss:
                            res["token"] = "pass"
                            res["cell"] = "PASS"
                        else:
                            res["token"] = "fail"
                            res["cell"] = "Lỗi: " + ",".join(f"BT{n}" for n in miss)
                    elif kind == "skip":
                        res["token"] = "skip"
                        res["cell"] = "Bỏ qua"
                    else:  # stop
                        res["token"] = "stop"
                        res["cell"] = None
                    ev.set()
                try:
                    win.destroy()
                except Exception:
                    pass

            tk.Button(bf, text="Xác nhận", bg="#16a34a", fg="white",
                      activebackground="#16a34a", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("confirm")).pack(side="left", padx=8)
            tk.Button(bf, text="Bỏ qua", bg="#64748b", fg="white",
                      activebackground="#64748b", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("skip")).pack(side="left", padx=8)
            win.protocol("WM_DELETE_WINDOW", lambda: finish("skip"))
            self._dialog_place(win)

            def tick():
                if ev.is_set():
                    return
                if self._stop_test.is_set():
                    finish("stop")
                    return
                marked = set(self._bt_test_pressed) | manual
                for n in marked:
                    if n in cells:
                        d, o = cells[n]
                        d.itemconfig(o, fill="#22c55e")
                if len(marked) >= N_RELAY:
                    finish("auto")
                    return
                win.after(150, tick)
            tick()

        self.after(0, show)
        ev.wait()
        self._bt_test_active = False
        return (res["token"], res["cell"])

    def _run_dip_test(self):
        # Lấy giá trị DIP hiện tại làm mốc rồi theo dõi từng bit bị gạt (giống BT)
        self.worker.clear_rx()
        resp = self.worker.send_cmd("dip", 0.4)
        m = RX_DIP_RE.search(resp)
        if m:
            self._dip_last_value = int(m.group(1), 16)
            self.after(0, lambda v=self._dip_last_value: self._update_dip(v))
        self._dip_test_seen = set()
        self._dip_test_active = True
        res = {"token": None, "cell": None}
        ev = threading.Event()

        def show():
            manual = set()
            win = tk.Toplevel(self)
            win.title("Kiểm tra DIP Switch")
            win.configure(bg=WHITE)
            win.resizable(False, False)
            tk.Label(win, text="Gạt lần lượt DIP1 -> DIP4 trên thiết bị\n"
                               "(gạt sẽ chuyển xanh). Đủ 4 bit sẽ tự động PASS.\n"
                               "Hoặc click vào ô để đánh dấu thủ công.",
                     bg=WHITE, fg="#1e293b", font=("Segoe UI", 11),
                     justify="left", padx=24).pack(pady=(16, 10))
            statf = tk.Frame(win, bg=WHITE)
            statf.pack(pady=4)
            cells = {}

            def manual_mark(n):
                manual.add(n)

            for n in range(1, 5):
                cell = tk.Frame(statf, bg=WHITE, cursor="hand2")
                cell.pack(side="left", padx=12)
                dot = tk.Canvas(cell, width=20, height=20, bg=WHITE,
                                highlightthickness=0)
                oid = dot.create_oval(2, 2, 18, 18, fill="#cbd5e1", outline="")
                dot.pack()
                lbl = tk.Label(cell, text=f"DIP{n}", bg=WHITE,
                               font=("Segoe UI", 9))
                lbl.pack()
                cells[n] = (dot, oid)
                for w in (cell, dot, lbl):
                    w.bind("<Button-1>", lambda e, nn=n: manual_mark(nn))

            bf = tk.Frame(win, bg=WHITE)
            bf.pack(pady=(10, 16))

            def finish(kind):
                if not ev.is_set():
                    marked = set(self._dip_test_seen) | manual
                    if kind in ("confirm", "auto"):
                        miss = [n for n in range(1, 5) if n not in marked]
                        if kind == "auto" or not miss:
                            res["token"] = "pass"
                            res["cell"] = "PASS"
                        else:
                            res["token"] = "fail"
                            res["cell"] = "Lỗi: " + ",".join(f"DIP{n}" for n in miss)
                    elif kind == "skip":
                        res["token"] = "skip"
                        res["cell"] = "Bỏ qua"
                    else:  # stop
                        res["token"] = "stop"
                        res["cell"] = None
                    ev.set()
                try:
                    win.destroy()
                except Exception:
                    pass

            tk.Button(bf, text="Xác nhận", bg="#16a34a", fg="white",
                      activebackground="#16a34a", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("confirm")).pack(side="left", padx=8)
            tk.Button(bf, text="Bỏ qua", bg="#64748b", fg="white",
                      activebackground="#64748b", font=("Segoe UI", 10, "bold"),
                      relief="flat", bd=0, padx=18, pady=8,
                      command=lambda: finish("skip")).pack(side="left", padx=8)
            win.protocol("WM_DELETE_WINDOW", lambda: finish("skip"))
            self._dialog_place(win)

            def tick():
                if ev.is_set():
                    return
                if self._stop_test.is_set():
                    finish("stop")
                    return
                marked = set(self._dip_test_seen) | manual
                for n in marked:
                    if n in cells:
                        d, o = cells[n]
                        d.itemconfig(o, fill="#22c55e")
                if len(marked) >= 4:
                    finish("auto")
                    return
                win.after(150, tick)
            tick()

        self.after(0, show)
        ev.wait()
        self._dip_test_active = False
        return (res["token"], res["cell"])

    def _run_rs485_test(self, text):
        # Gửi text qua RS485; với loopback (A-B / TX-RX) thiết bị sẽ vọng text
        # ve console (rs485_process forward). Kiem tra text co trong phan hoi.
        self.worker.clear_rx()
        resp = self.worker.send_cmd(f"rs485 {text}", 1.2)
        if text in resp:
            return ("pass", "PASS")
        ans = self._ask_user(
            "Test RS485",
            f"Đã gửi '{text}' qua RS485 nhưng chưa nhận được vòng về.\n"
            "Kiểm tra dây loopback RS485 (A-B hoặc TX-RX).\n\n"
            "Thử lại, bỏ qua bước này, hay đánh FAIL?",
            [("Thử lại", "retry", BLUE_BTN),
             ("Bỏ qua", "skip", "#64748b"),
             ("FAIL", "fail", "#ef4444")])
        if ans == "retry":
            return self._run_rs485_test(text)
        if ans == "fail":
            return ("fail", "FAIL")
        if ans == "stop":
            return ("stop", None)
        return ("skip", "Bo qua")

    def _save_report(self, fw, cells, n_pass, n_fail, n_total, stopped):
        header = ["Serial", "Thoi gian", "COM", "Baud", "FW",
                  *[t["name"] for t in TESTS], "PASS/Tong", "Ket luan"]
        ket_luan = "DUNG GIUA" if stopped else ("OK" if n_fail == 0 else "LOI")
        row = [self._serial_no, time.strftime("%Y-%m-%d %H:%M:%S"),
               self._sel_port(), self.cbo_baud.get(), fw,
               *cells, f"{n_pass}/{n_total}", ket_luan]
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            import csv
            path = report_path()[:-5] + ".csv"
            try:
                is_new = not os.path.exists(path)
                with open(path, "a", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    if is_new:
                        w.writerow(header)
                    w.writerow(row)
                self._log(f"[Không có openpyxl - đã ghi CSV: {path}]\n", "tx")
            except PermissionError:
                messagebox.showerror(
                    APP_TITLE,
                    "Không ghi được CSV - file đang mở?\nĐóng file rồi chạy lại.")
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"Lỗi lưu report:\n{e}")
            return

        path = report_path()
        try:
            if os.path.exists(path):
                wb = load_workbook(path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Test report"
                ws.append(header)
            ws.append(row)
            wb.save(path)
            self._log(f"[Đã ghi report: {path}]\n", "tx")
        except PermissionError:
            messagebox.showerror(
                APP_TITLE,
                f"Không ghi được {REPORT_NAME} - file đang mở trong Excel?\n"
                "Đóng file rồi chạy test lại.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Lỗi lưu report:\n{e}")

    def _set_row(self, idx, text, tag):
        self.after(0, lambda: self.tree.item(str(idx), values=(text,), tags=(tag,)))

    def _on_close(self):
        self.worker.close()
        self.destroy()


if __name__ == "__main__":
    App().mainloop()
